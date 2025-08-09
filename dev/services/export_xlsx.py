# app/services/export_xlsx.py
from __future__ import annotations
from datetime import datetime, date
from typing import Iterable, Tuple
import os
import tempfile

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, numbers

from app.models.operation import Operation
from app.repo.records import aggregate_by_category

def _auto_width(ws) -> None:
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            val = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(val) + 1)
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(w, 60)

def _format_money(cell) -> None:
    cell.number_format = numbers.FORMAT_NUMBER_00

def _title(ws, text: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")

def build_xlsx(
    ops: Iterable[Operation],
    start: date,
    end: date,
    user_label: str = "",
) -> str:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Операции"

    _title(ws1, f"Операции {start.isoformat()} — {end.isoformat()} {user_label}".strip())

    headers = ["#", "Дата/время", "Тип", "Сумма BYN", "Категория", "Описание"]
    ws1.append(headers)
    for i, h in enumerate(headers, 1):
        ws1.cell(row=2, column=i).font = Font(bold=True)

    row_idx = 3
    for idx, o in enumerate(ops, 1):
        ws1.cell(row=row_idx, column=1, value=idx)
        ws1.cell(row=row_idx, column=2, value=o.created_at.strftime("%Y-%m-%d %H:%M"))
        ws1.cell(row=row_idx, column=3, value=("Расход" if o.type == "expense" else "Доход"))
        c = ws1.cell(row=row_idx, column=4, value=float(abs(o.amount)))
        _format_money(c)
        ws1.cell(row=row_idx, column=5, value=o.category or "")
        ws1.cell(row=row_idx, column=6, value=o.description or "")
        row_idx += 1

    _auto_width(ws1)

    # Сводка
    ws2 = wb.create_sheet(title="Сводка")
    _title(ws2, "Сводка по категориям")

    headers2 = ["Категория", "Сумма BYN (знак: + доход, - расход)"]
    ws2.append(headers2)
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i).font = Font(bold=True)

    agg = {}
    # нужен знак: доход +, расход -
    for o in ops:
        sign = 1.0 if o.type == "income" else -1.0
        agg[o.category or "Прочее"] = agg.get(o.category or "Прочее", 0.0) + sign * float(abs(o.amount))

    row_idx = 3
    for cat, val in sorted(agg.items()):
        ws2.cell(row=row_idx, column=1, value=cat)
        c = ws2.cell(row=row_idx, column=2, value=round(val, 2))
        _format_money(c)
        row_idx += 1

    _auto_width(ws2)

    # Сохраняем во временный файл
    fd, path = tempfile.mkstemp(prefix="fin_export_", suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path
